import os
import io
import zipfile
import datetime
import uuid
import subprocess  # NOWOŚĆ: Do uruchamiania LibreOffice
import shutil  # Do operacji na plikach
import time
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from collections import defaultdict
from functools import wraps
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import click



# ==============================================================================
# 1. KONFIGURACJA APLIKACJI
# ==============================================================================
app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['SECRET_KEY'] = 'twoj-bardzo-tajny-i-dlugi-klucz-zmien-go-koniecznie'
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(BASE_DIR, 'instance', 'app.db')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# --- UPLOAD FOLDER (Zostaje w projekcie) ---
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'user_uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# --- TEMPORARY FILES DIR (KLUCZOWA ZMIANA ŚCIEŻKI) ---
# Ustawienie globalnej, prostej ścieżki bez spacji.
# MUSISZ ręcznie utworzyć ten folder na dysku C:
TEMP_FILES_DIR = "C:\\temp_certs"
os.makedirs(TEMP_FILES_DIR, exist_ok=True) # Zapewnienie, że folder istnieje

# --- KONFIGURACJA LIBREOFFICE (Ścieżka na Windows) ---
LIBREOFFICE_PATH = "C:\\Program Files\\LibreOffice\\program\\soffice.exe"

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = "Proszę się zalogować, aby uzyskać dostęp do tej strony."
login_manager.login_message_category = "info"

# ==============================================================================
# 2. MODELE BAZY DANYCH
# ==============================================================================

template_assignments = db.Table('template_assignments',
                                db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
                                db.Column('template_id', db.Integer, db.ForeignKey('template.id'), primary_key=True)
                                )


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='user')

    company_name = db.Column(db.String(150), nullable=True)
    email = db.Column(db.String(120), nullable=True)
    phone = db.Column(db.String(30), nullable=True)
    contact_person = db.Column(db.String(150), nullable=True)
    nip = db.Column(db.String(20), nullable=True)

    assigned_templates = db.relationship('Template', secondary=template_assignments, lazy='dynamic',
                                         back_populates='assigned_users')
    generation_logs = db.relationship('GenerationLog', backref='user', lazy='dynamic', cascade="all, delete-orphan")

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    @property
    def is_admin(self):
        return self.role == 'admin'


class Template(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    numbering_format = db.Column(db.String(255), nullable=True)

    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    creator = db.relationship('User', backref='created_templates')

    mappings = db.relationship('CellMapping', backref='template', lazy='dynamic', cascade="all, delete-orphan")
    assigned_users = db.relationship('User', secondary=template_assignments, lazy='dynamic',
                                     back_populates='assigned_templates')
    generation_logs = db.relationship('GenerationLog', backref='template', lazy='dynamic', cascade="all, delete-orphan")

    @property
    def usage_count(self):
        return self.generation_logs.count()

    @property
    def total_certificates_generated(self):
        logs = self.generation_logs.all()
        return sum(log.certificate_count for log in logs)


class CellMapping(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    template_id = db.Column(db.Integer, db.ForeignKey('template.id'), nullable=False)
    field_name = db.Column(db.String(80), nullable=False)
    cell_address = db.Column(db.String(10), nullable=False)
    is_required = db.Column(db.Boolean, default=False, nullable=False)


class GenerationLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    generation_date = db.Column(db.DateTime, nullable=False, default=datetime.datetime.now)  # Czas lokalny
    certificate_count = db.Column(db.Integer, nullable=False)

    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    template_id = db.Column(db.Integer, db.ForeignKey('template.id'), nullable=False)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# ==============================================================================
# 3. FUNKCJE POMOCNICZE (NOWOŚĆ - KONWERSJA PDF)
# ==============================================================================

def convert_to_pdf(source_xlsx_path, output_folder):
    """
    Konwertuje plik XLSX na PDF przy użyciu LibreOffice.
    Wersja ostateczna, zoptymalizowana dla serwera Linux.
    """
    # Na serwerach Linux/Heroku wystarczy komenda 'libreoffice'
    LIBREOFFICE_PATH = "/app/vendor/libreoffice/instdir/program/soffice"

    command = [
        LIBREOFFICE_PATH,
        '--headless',
        '--nologo',
        '--convert-to', 'pdf:calc_pdf_Export',  # Używamy stabilnego filtru dla arkuszy
        '--outdir', output_folder,
        source_xlsx_path
    ]

    print(f"Wywoływanie LibreOffice (Linux): {' '.join(command)}")

    try:
        # Ten kod zadziała na Linuxie. Na Windows może nadal nie działać.
        process = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=45)

        if process.returncode != 0:
            error_output = process.stderr.decode(errors='ignore')
            raise Exception(f"Błąd konwersji (Kod: {process.returncode}). Błąd LibreOffice/Linux: {error_output}")

    except Exception as e:
        # Złapanie błędu w środowisku Linux
        raise Exception(f"Krytyczny błąd konwersji PDF: {e}")

    # Sprawdzanie i zwracanie ścieżki do PDF
    base_name = os.path.splitext(os.path.basename(source_xlsx_path))[0]
    pdf_filename = f"{base_name}.pdf"
    pdf_path = os.path.join(output_folder, pdf_filename)

    if not os.path.exists(pdf_path):
        raise Exception("Konwersja zakończona, ale plik PDF nie został utworzony. Prawdopodobnie błąd renderowania.")

    return pdf_path

# ==============================================================================
# 4. DEKORATORY I CLI
# ==============================================================================


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("Dostęp do tej strony wymaga uprawnień administratora.", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


@app.cli.command("init-db")
def init_db_command():
    with app.app_context():
        db.drop_all()
        db.create_all()
    click.echo("Zainicjalizowano bazę danych.")


@app.cli.command("create-admin")
@click.argument("username")
@click.argument("password")
def create_admin(username, password):
    with app.app_context():
        if User.query.filter_by(username=username).first():
            print(f"Użytkownik '{username}' już istnieje.")
            return
        new_admin = User(username=username, role='admin')
        new_admin.set_password(password)
        db.session.add(new_admin)
        db.session.commit()
        print(f"Administrator '{username}' został pomyślnie utworzony.")


# ==============================================================================
# 5. TRASY (ROUTES)
# ==============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            login_user(user, remember=True)
            if user.is_admin:
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        else:
            flash('Nieprawidłowy login lub hasło.', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Zostałeś pomyślnie wylogowany.', 'success')
    return redirect(url_for('login'))


@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard'))
    user_templates = current_user.assigned_templates.order_by(Template.name).all()
    return render_template('dashboard.html', templates=user_templates)


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    if request.method == 'POST':
        current_user.contact_person = request.form.get('contact_person')
        current_user.phone = request.form.get('phone')
        current_user.email = request.form.get('email')

        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if new_password:
            if new_password != confirm_password:
                flash("Podane hasła nie są identyczne!", "danger")
                return redirect(url_for('user_profile'))
            current_user.set_password(new_password)
            flash("Twoje hasło zostało zmienione.", "success")

        db.session.commit()
        flash("Dane profilowe zostały zaktualizowane.", "info")
        return redirect(url_for('user_profile'))

    return render_template('profile.html', user=current_user)


# --- ADMINISTRACJA ---

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    all_templates = Template.query.order_by(Template.name).all()
    fields_to_map = ['imie_nazwisko', 'data_urodzenia', 'miejsce_urodzenia', 'pesel', 'instruktor', 'numer_certyfikatu',
                     'data_szkolenia', 'waznosc']
    return render_template('admin_dashboard.html', templates=all_templates, fields_to_map=fields_to_map)


@app.route('/admin/users')
@login_required
@admin_required
def manage_users():
    users = User.query.order_by(User.username).all()
    return render_template('manage_users.html', users=users)


@app.route('/admin/user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def user_details(user_id):
    user = db.session.get(User, user_id)
    if not user or user.is_admin:
        flash("Nie znaleziono użytkownika lub próbujesz edytować admina.", "danger")
        return redirect(url_for('manage_users'))

    if request.method == 'POST':
        user.company_name = request.form.get('company_name')
        user.email = request.form.get('email')
        user.phone = request.form.get('phone')
        user.contact_person = request.form.get('contact_person')
        user.nip = request.form.get('nip')

        new_password = request.form.get('new_password')
        if new_password:
            user.set_password(new_password)
            flash("Hasło zostało zmienione.", "info")

        db.session.commit()
        flash("Dane użytkownika zostały zaktualizowane.", "success")
        return redirect(url_for('user_details', user_id=user.id))

    logs = getattr(user, 'generation_logs', [])
    if hasattr(logs, 'order_by'):
        logs = logs.order_by(GenerationLog.generation_date.desc()).all()

    assigned_templates = user.assigned_templates.order_by(Template.name).all()
    return render_template('user_details.html', user=user, logs=logs, assigned_templates=assigned_templates)


@app.route('/admin/users/add', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    if not username or not password:
        flash('Nazwa użytkownika i hasło są wymagane.', 'warning')
        return redirect(url_for('manage_users'))
    if User.query.filter_by(username=username).first():
        flash('Ta nazwa użytkownika jest już zajęta.', 'danger')
        return redirect(url_for('manage_users'))
    new_user = User(username=username, role='user')
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    flash(f'Użytkownik "{username}" został pomyślnie dodany.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    user_to_delete = db.session.get(User, user_id)
    if not user_to_delete:
        flash("Nie znaleziono użytkownika.", 'danger')
        return redirect(url_for('manage_users'))
    if user_to_delete.is_admin:
        flash("Nie można usunąć konta administratora.", 'danger')
        return redirect(url_for('manage_users'))
    db.session.delete(user_to_delete)
    db.session.commit()
    flash(f'Użytkownik "{user_to_delete.username}" został usunięty.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/admin/template/add', methods=['POST'])
@login_required
@admin_required
def add_template():
    template_file = request.files.get('template_file')
    template_name = request.form.get('template_name')
    numbering_format = request.form.get('numbering_format')

    if not template_name or not template_file or not request.form.get('mapping_imie_nazwisko'):
        flash('Nazwa szablonu, plik oraz mapowanie dla "Imię i Nazwisko" są obowiązkowe.', 'danger')
        return redirect(url_for('admin_dashboard'))

    mappings_data = []
    for key, value in request.form.items():
        if key.startswith('mapping_') and value:
            field_name = key.replace('mapping_', '')
            is_req = request.form.get(f'is_required_{field_name}') == 'on'
            mappings_data.append({'field_name': field_name, 'cell_address': value.upper(), 'is_required': is_req})

    custom_field_index = 0
    while True:
        custom_name_key = f'custom_name_{custom_field_index}'
        if custom_name_key in request.form:
            name = request.form.get(custom_name_key)
            cell = request.form.get(f'custom_cell_{custom_field_index}')
            is_req = request.form.get(f'custom_is_required_{custom_field_index}') == 'on'
            if name and cell:
                mappings_data.append({'field_name': f'custom_{name.replace(" ", "_")}', 'cell_address': cell.upper(),
                                      'is_required': is_req})
            custom_field_index += 1
        else:
            break

    filename = secure_filename(f"{current_user.id}_{uuid.uuid4().hex[:8]}_{template_file.filename}")
    template_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

    new_template = Template(name=template_name, filename=filename, creator=current_user,
                            numbering_format=numbering_format)
    db.session.add(new_template)
    for mapping in mappings_data:
        new_template.mappings.append(CellMapping(**mapping))
    db.session.commit()
    flash('Nowy szablon został pomyślnie dodany!', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/template/<int:template_id>/download')
@login_required
@admin_required
def download_template_source(template_id):
    template = db.session.get(Template, template_id)
    if not template:
        flash("Nie znaleziono szablonu.", "danger")
        return redirect(url_for('admin_dashboard'))
    try:
        return send_file(
            os.path.join(app.config['UPLOAD_FOLDER'], template.filename),
            as_attachment=True,
            download_name=template.filename
        )
    except Exception as e:
        print(f"Błąd pobierania pliku szablonu: {e}")
        flash("Wystąpił błąd podczas pobierania pliku. Plik może nie istnieć.", "danger")
        return redirect(url_for('edit_template', template_id=template.id))


@app.route('/admin/template/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_template(template_id):
    template = db.session.get(Template, template_id)
    if not template:
        flash("Nie znaleziono szablonu.", 'danger')
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        template.name = request.form.get('template_name')
        template.numbering_format = request.form.get('numbering_format')

        template.mappings.delete()
        mappings_data = []
        for key, value in request.form.items():
            if key.startswith('mapping_') and value:
                field_name = key.replace('mapping_', '')
                is_req = request.form.get(f'is_required_{field_name}') == 'on'
                mappings_data.append({'field_name': field_name, 'cell_address': value.upper(), 'is_required': is_req})

        custom_field_index = 0
        while True:
            custom_name_key = f'custom_name_{custom_field_index}'
            if custom_name_key in request.form:
                name = request.form.get(custom_name_key)
                cell = request.form.get(f'custom_cell_{custom_field_index}')
                is_req = request.form.get(f'custom_is_required_{custom_field_index}') == 'on'
                if name and cell:
                    mappings_data.append(
                        {'field_name': f'custom_{name.replace(" ", "_")}', 'cell_address': cell.upper(),
                         'is_required': is_req})
                custom_field_index += 1
            else:
                break

        for mapping in mappings_data:
            template.mappings.append(CellMapping(**mapping))

        assigned_user_ids = request.form.getlist('assigned_users')
        template.assigned_users = User.query.filter(User.id.in_(assigned_user_ids)).all()

        db.session.commit()
        flash(f'Szablon "{template.name}" został zaktualizowany.', 'success')
        return redirect(url_for('admin_dashboard'))

    all_users = User.query.filter_by(role='user').order_by(User.username).all()
    assigned_user_ids = {user.id for user in template.assigned_users}
    fields_to_map = ['imie_nazwisko', 'data_urodzenia', 'miejsce_urodzenia', 'pesel', 'instruktor', 'numer_certyfikatu',
                     'data_szkolenia', 'waznosc']
    mappings_dict = {m.field_name: {'cell': m.cell_address, 'required': m.is_required} for m in template.mappings}
    custom_mappings = {k: v for k, v in mappings_dict.items() if k.startswith('custom_')}

    return render_template('edit_template.html',
                           template=template,
                           fields_to_map=fields_to_map,
                           mappings_dict=mappings_dict,
                           custom_mappings=custom_mappings,
                           all_users=all_users,
                           assigned_user_ids=assigned_user_ids)


@app.route('/admin/template/<int:template_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_template(template_id):
    template = db.session.get(Template, template_id)
    if template:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], template.filename))
        except OSError as e:
            print(f"Błąd podczas usuwania pliku szablonu: {e}")
        db.session.delete(template)
        db.session.commit()
        flash(f'Szablon "{template.name}" został usunięty.', 'success')
    return redirect(url_for('admin_dashboard'))


# --- API (Frontend) ---

@app.route('/api/template/<int:template_id>')
@login_required
def get_template_details(template_id):
    if current_user.is_admin:
        template = db.session.get(Template, template_id)
    else:
        template = current_user.assigned_templates.filter_by(id=template_id).first()
    if not template:
        return jsonify({'error': 'Template not found or not assigned'}), 404
    mapped_fields = [{'name': m.field_name, 'required': m.is_required} for m in template.mappings]
    return jsonify({'mapped_fields': mapped_fields, 'numbering_format': template.numbering_format})


# ==============================================================================
# GŁÓWNA FUNKCJA GENEROWANIA (SILNIK LIBREOFFICE + OPENPYXL)
# ==============================================================================
@app.route('/api/generate', methods=['POST'])
@login_required
def generate_certificates():
    # KROK 1: Koniec z pythoncom/win32com. Teraz OpenPyXL + LibreOffice.
    try:
        participant_list_file = request.files.get('participant_list')
        template_id = request.form.get('template_id')

        if not participant_list_file or not template_id:
            return jsonify({'error': 'Brak pliku z listą lub nie wybrano szablonu.'}), 400

        if current_user.is_admin:
            template = db.session.get(Template, template_id)
        else:
            template = current_user.assigned_templates.filter_by(id=template_id).first()

        if not template:
            return jsonify({'error': 'Nie znaleziono szablonu.'}), 404

        # Mapowanie pól
        cell_mappings = {m.field_name: m.cell_address for m in template.mappings}

        # Ładowanie listy uczestników
        wb_list = load_workbook(participant_list_file)
        rows_to_process = [row for row in wb_list.active.iter_rows(min_row=2, max_col=4, values_only=True) if
                           row and row[0]]

        # Przygotowanie ZIP
        zip_buffer = io.BytesIO()

        # Ścieżka do wzorca
        template_path = os.path.join(app.config['UPLOAD_FOLDER'], template.filename)

        try:
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:

                for i, row in enumerate(rows_to_process):
                    # 1. Przygotowanie danych (tak samo jak wcześniej)
                    data_to_insert = {
                        'imie_nazwisko': row[0] if len(row) > 0 and row[0] else "",
                        'data_urodzenia': str(row[1]) if len(row) > 1 and row[1] else "",
                        'miejsce_urodzenia': str(row[2]) if len(row) > 2 and row[2] else "",
                        'pesel': str(int(row[3])) if len(row) > 3 and row[3] else "",
                    }

                    # Daty i numeracja (logika biznesowa)
                    training_date_str = request.form.get('training_date', '')
                    if training_date_str:
                        dt_object = datetime.datetime.strptime(training_date_str, "%Y-%m-%d")
                        data_to_insert['data_szkolenia'] = dt_object.strftime('%d.%m.%Y')

                        if template.numbering_format:
                            start_number = int(request.form.get('start_number', 1))
                            replacements = {
                                '{NUMER}': str(start_number + i),
                                '{NUMER_STARTOWY}': str(start_number),
                                '{ROK}': str(dt_object.year),
                                '{RR}': dt_object.strftime('%y'),
                                '{MIESIAC}': dt_object.strftime('%m'),
                                '{DZIEN}': dt_object.strftime('%d'),
                                '{ID_UZYTKOWNIKA}': str(current_user.id),
                                '{LOGIN_UZYTKOWNIKA}': current_user.username
                            }
                            cert_number = template.numbering_format
                            for placeholder, value in replacements.items():
                                cert_number = cert_number.replace(placeholder, value)
                            data_to_insert['numer_certyfikatu'] = cert_number
                        else:
                            start_number = int(request.form.get('start_number', 1))
                            data_to_insert[
                                'numer_certyfikatu'] = f"{start_number + i}/{dt_object.month}/{dt_object.year}"

                    data_to_insert['instruktor'] = request.form.get('instructor_name', '')

                    # Ważność
                    validity_type = request.form.get('validity_type')
                    if validity_type == 'years':
                        validity_years = int(request.form.get('validity_value_years', 1))
                        if validity_years == 1:
                            data_to_insert['waznosc'] = "1 rok od daty wystawienia dokumentu"
                        else:
                            data_to_insert['waznosc'] = f"{validity_years} lata od daty wystawienia dokumentu"
                    elif validity_type == 'date':
                        validity_date_str = request.form.get('validity_value_date')
                        if validity_date_str:
                            data_to_insert['waznosc'] = datetime.datetime.strptime(validity_date_str,
                                                                                   '%Y-%m-%d').strftime('%d.%m.%Y')
                        else:
                            data_to_insert['waznosc'] = "Bezterminowo"
                    else:
                        data_to_insert['waznosc'] = "Bezterminowo"

                    for field_name in cell_mappings.keys():
                        if field_name.startswith('custom_'):
                            form_key = f"custom_field_{field_name.replace('custom_', '')}"
                            data_to_insert[field_name] = request.form.get(form_key, '')

                    # 2. EDYCJA EXCELA (OpenPyXL zamiast COM)
                    # Ładujemy szablon do pamięci dla KAŻDEGO uczestnika, żeby mieć czystą kopię
                    wb_cert = load_workbook(template_path)
                    ws_cert = wb_cert.active  # Zakładamy, że dane są na 1. arkuszu

                    # Grupujemy dane, jeśli wiele pól ma ten sam adres
                    final_cell_data = defaultdict(list)
                    for field, value in data_to_insert.items():
                        if field in cell_mappings and value:
                            cell_addr = cell_mappings[field]
                            final_cell_data[cell_addr].append(str(value))

                    # Wpisujemy dane do komórek
                    for cell_addr, values_list in final_cell_data.items():
                        final_value = ", ".join(values_list)
                        # OpenPyXL używa zapisu: ws['A1'] = 'wartość'
                        ws_cert[cell_addr] = final_value

                    # 3. ZAPIS TYMCZASOWY I KONWERSJA
                    base_filename = f"{data_to_insert.get('numer_certyfikatu', 'cert').replace('/', '_')}_{uuid.uuid4().hex[:4]}"
                    temp_xlsx_path = os.path.join(TEMP_FILES_DIR, f"{base_filename}.xlsx")

                    # Zapisujemy wypełnionego Excela
                    wb_cert.save(temp_xlsx_path)
                    wb_cert.close()

                    # Konwersja na PDF (LibreOffice)
                    try:
                        pdf_path = convert_to_pdf(temp_xlsx_path, TEMP_FILES_DIR)

                        # Dodajemy do ZIP
                        zf.write(pdf_path, arcname=f"{base_filename.split('_')[0]}_{i + 1}.pdf")
                        # Opcjonalnie: dodajemy też Excela (jeśli chcesz)
                        # zf.write(temp_xlsx_path, arcname=f"{base_filename}.xlsx")

                        # Sprzątanie pojedynczych plików
                        os.remove(pdf_path)
                        os.remove(temp_xlsx_path)

                    except Exception as conversion_error:
                        print(f"Błąd konwersji dla {data_to_insert['imie_nazwisko']}: {conversion_error}")
                        # W razie błędu, dodajemy chociaż Excela do ZIP
                        zf.write(temp_xlsx_path, arcname=f"BLAD_KONWERSJI_{base_filename}.xlsx")
                        os.remove(temp_xlsx_path)

            # 4. ZAPIS LOGÓW (Tylko jeśli ZIP utworzony pomyślnie)
            try:
                cert_count = len(rows_to_process)
                new_log = GenerationLog(
                    certificate_count=cert_count,
                    user_id=current_user.id,
                    template_id=template.id,
                    generation_date=datetime.datetime.now()
                )
                db.session.add(new_log)
                db.session.commit()
            except Exception as log_error:
                print(f"Błąd podczas zapisywania logów: {log_error}")

            zip_buffer.seek(0)
            return send_file(zip_buffer, mimetype='application/zip', as_attachment=True,
                             download_name='certyfikaty.zip')

        except Exception as e:
            print(f"Błąd ZIP/Pętli: {e}")
            raise e

    except Exception as e:
        print(f"KRYTYCZNY BŁĄD SERWERA: {e}")
        return jsonify({'error': f'Błąd generowania: {str(e)}'}), 500


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5001)